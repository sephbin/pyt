from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill, Alignment
import csv

filepath = r"D:\Dropbox (COX Architecture)\Rugby Australia - Phase 1 Sports Architect\Phase 2 - RA Cox RWC Bid Documentation\02 Ticketing Maps\Suncorp Stadium Brisbane (003).xlsx"
wb = load_workbook(filepath, data_only=True)

seats = []

# for ws in wb.worksheets[3:]:
for ws in wb.worksheets:
	title = ws.title
	# print(title)
	maxCell = 0
	for topCellIndex, topCell in enumerate(list(ws.iter_rows())[10]):
		# print(topCellIndex, topCell)
		if topCellIndex > 5 and topCell.value == None:
			maxCell = topCellIndex
			# print(maxCell)
			break

	aisleNumber = 1
	for rowIndex, row in enumerate(ws.iter_rows()):
		# print("!!!!!!!!!!!",row[3].value,row[4].value)
		if rowIndex > 10 and row[4].value == None:
			break
		if rowIndex > 10:
			seatNumber = 1
			for cellIndex, cell in enumerate(row):
				if cellIndex == maxCell:
					break
				if cellIndex > 5:
					if cell.value:
						apOb = {
						"seatNumber": seatNumber,
						"aisleNumber": aisleNumber,
						"section": title,
						"seatType": cell.value
						}
						seats.append(apOb)s
					seatNumber += 1
		aisleNumber += 1

csv_columns = list(seats[0])
csv_file = r"D:\Dropbox (COX Architecture)\Rugby Australia - Phase 1 Sports Architect\Phase 2 - RA Cox RWC Bid Documentation\02 Ticketing Maps\BNE seats.csv"
try:
	with open(csv_file, 'w',newline='') as csvfile:
		writer = csv.DictWriter(csvfile, fieldnames=csv_columns)
		writer.writeheader()
		for d in seats:
			writer.writerow(d)
except Exception as e:
	print(e)