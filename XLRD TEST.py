import xlrd
from openpyxl import Workbook
import json
from xlrd.sheet import ctype_text
# Open the workbook
originalWorkbook = "D:\\Users\\s-abutler\\Downloads\\SFS.xlsx"
newWorkbook = "D:\\Users\\s-abutler\\Downloads\\200805 SFS DUMP.xlsx"

original_Data = []
new_Data = []

indexID = ["Level","Room: Number","Family","Type"]


original_workbook = xlrd.open_workbook(originalWorkbook)
original_sheet_names = original_workbook.sheet_names()
original_sheet = original_workbook.sheet_by_name(original_sheet_names[0])
original_sheet = original_workbook.sheet_by_index(0)
original_row = original_sheet.row(1)
original_num_cols = original_sheet.ncols   # Number of columns
for row_idx in range(1, original_sheet.nrows):    # Iterate through rows
	apOB = {}
	for col_idx in range(0, original_num_cols):  # Iterate through columns
		cell_obj = original_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
		key = original_sheet.cell(0, col_idx)
		apOB[key.value]= original_sheet.cell(row_idx, col_idx).value
		# print ('Column: [%s] cell_obj: [%s]' % (col_idx, cell_obj))
	original_Data.append(apOB)

new_workbook = xlrd.open_workbook(newWorkbook)
new_sheet_names = new_workbook.sheet_names()
new_sheet = new_workbook.sheet_by_name(new_sheet_names[0])
new_sheet = new_workbook.sheet_by_index(0)
new_row = new_sheet.row(1)
new_num_cols = new_sheet.ncols   # Number of columns
for row_idx in range(1, new_sheet.nrows):    # Iterate through rows
	apOB = {}
	for col_idx in range(0, new_num_cols):  # Iterate through columns
		cell_obj = new_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
		key = new_sheet.cell(0, col_idx)
		apOB[key.value]= new_sheet.cell(row_idx, col_idx).value
		# print ('Column: [%s] cell_obj: [%s]' % (col_idx, cell_obj))
	new_Data.append(apOB)


# print(original_Data)
# print(new_Data)


for oA in new_Data:
	indexNotFound = True

	for oR in original_Data:
		oAIndex = ""
		oRIndex = ""
		for indexKey in indexID:
			oAIndex += str(oA[indexKey])+"`"
			oRIndex += str(oR[indexKey])+"`"
		if oAIndex == oRIndex:
			indexNotFound = False
			hasChanged = oA != oR
			
			changedKeys = []
			if hasChanged:
				changedLog = []
				for k, v in oA.items():

					try:
						print(k, oR[k], v)
						if oR[k] != v:
							fromVal = oR[k]
							toVal = v
							if fromVal == "":
								fromVal = "\"\""
							if toVal == "":
								toVal = "\"\""
							changedLog.append(f"{k}: {fromVal} to {toVal}")
							changedKeys.append(k)
					except:
						pass
				changedLog = ";\n".join(changedLog)
				oA["change log"] = changedLog
			else:
				oA["change log"] = ""
			oA["changedKeys"] = changedKeys
			oA["hasChanged"] = hasChanged
	if indexNotFound:
		oA["hasChanged"] = True
		oA["change log"] = "New"
		oA["changedKeys"] = []


for oR in original_Data:
	indexNotFound = True
	for oA in new_Data:
		oAIndex = ""
		oRIndex = ""
		for indexKey in indexID:
			oAIndex += str(oA[indexKey])+"`"
			oRIndex += str(oR[indexKey])+"`"
		if oAIndex == oRIndex:
			indexNotFound = False

	if indexNotFound:
		oR["hasChanged"] = True
		oR["change log"] = "Deleted"
		oR["changedKeys"] = []
		new_Data.append(oR)


print(new_Data)

wb = Workbook()

# grab the active worksheet
ws = wb.active

ignoreHeaders = ["changedKeys"]
headers = []
col_index = 0
for  k in new_Data[0]:
	if k not in ignoreHeaders:
		headers.append(k)
		ws.cell(row=1, column=col_index+1, value=k)
		col_index += 1

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
alignment = Alignment(horizontal='general', vertical='bottom', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)

for row_index, row in enumerate(new_Data):
	print(row)
	for k, v in row.items():
		if k not in ignoreHeaders:
			col_index = headers.index(k)
			editCell = ws.cell(row=row_index+2, column=col_index+1)
			editCell.value = v

			editCell.alignment = Alignment(wrap_text=True, vertical='top',)
			# ws.cell(row=row_index+2, column=col_index+1).alignment.wrap_text = True
			if k in row["changedKeys"]:
				editCell.font = Font(color = "FF0000")
			if row["change log"]=="Deleted":
				editCell.font = Font(color = "FF0000", strike=True)
			if row["change log"]=="New":
				editCell.font = Font(color = "0000FF")
# Save the file
wb.save("SFS COMPARE.xlsx")