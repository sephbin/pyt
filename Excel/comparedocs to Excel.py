from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill, Alignment
# wb = Workbook()
# ws = wb.active
# ws.title = "New Title"
# # ws.sheet_properties.tabColor = "1072BA"
# c = ws.cell(row=4, column=2, value=10)
# c.font = Font(color="FF0000")
# c.fill = PatternFill("solid", fgColor="DDDDDD")

# print(ws.cell.__doc__)

# wb.save(r'D:\Users\s-abutler\Downloads\example.xlsx')

rev1 = "D"
rev2 = "E"

wb1 = load_workbook('\\\\sydsrv01\\projects\\2018\\218018.04\\06.0 SPECIFICATIONS & SCHEDULES\\6.2 Schedules\\Doors\\SFS-COX-01-SH-AR91XX02-%s.xlsx'%(rev1))
wb2 = load_workbook('\\\\sydsrv01\\projects\\2018\\218018.04\\06.0 SPECIFICATIONS & SCHEDULES\\6.2 Schedules\\Doors\\SFS-COX-01-SH-AR91XX02-%s.xlsx'%(rev2))
# wb2 = load_workbook(r'\\sydsrv01\projects\2018\218018.04\06.0 SPECIFICATIONS & SCHEDULES\6.2 Schedules\Doors\SFS-COX-01-SH-AR91XX02-E WIP.xlsx')
# wb2 = load_workbook(r'D:\Users\s-abutler\Documents\SFS-COX-01-SH-AR91XX02-E WIP.xlsx')
# print(wb2.worksheets)
# print(dir(wb2))
ws1 = wb1.worksheets[4]
ws2 = wb2.worksheets[4]

print(ws1,ws2)
# print(dir(ws))
# print( ws.min_column, ws.min_row)
# print(ws.max_column, ws.max_row)


headers1 = list(ws1.iter_rows(min_row=3, max_row=3))[0]
prevIndex = {}
ws1_rows = list(ws1.iter_rows(min_row=4, max_row=ws1.max_row))
for indR, row in enumerate(ws1_rows):
	for indC, (key,cell) in enumerate(zip(headers1, row)):
		k = key.value
		if k == "Door Number":
			prevIndex[cell.value] = indR


headers2 = list(ws2.iter_rows(min_row=3, max_row=3))[0]
ignoreChange = ["Room Name","Level"]
for indR, row in enumerate(ws2.iter_rows(min_row=4, max_row=ws2.max_row)):
	prevDict = {}
	changelog = []
	for indC, (key,cell) in enumerate(zip(headers2, row)):
		k = key.value
		v = cell.value
		if k == "Door Number":
			try:
				prevRow = ws1_rows[prevIndex[v]]
				for k1,v1 in zip(headers1, prevRow):
					# print(">>",k1.value,v1.value)
					try:	prevDict[k1.value] = v1.value
					except:	prevDict[k1.value] = None
				# print(prevDict)
			except Exception as e:
				pass
	# print(prevDict)
	# break
	error = True
	for indC, (key,cell) in enumerate(zip(headers2, row)):
		try:
			k = key.value
			v1 = prevDict[k]
			v2 = cell.value
			# print(k,v1,v2)
			if v1 != v2:
				cell.fill = PatternFill("solid", fgColor="FFFF00")
				if k not in ignoreChange:
					changelog.append("".join([k,": ",str(v1)," > ",str(v2)]))
			error = False
		except Exception as e:
			cell.fill = PatternFill("solid", fgColor="FFFFDD")
			# changelog.append("".join(["Added: ",key.value]))
	if error:
		changelog.append("New Door")
	cl = ws2.cell(row=4+indR, column=len(row)+1, value="\n".join(changelog))
	cl.alignment = Alignment(wrapText=True)
	# print(indR, )
	# print(dir(row))
	# print(row.count.__doc__)
	# print(changelog)
	# break
wb2.save('\\\\sydsrv01\\projects\\2018\\218018.04\\06.0 SPECIFICATIONS & SCHEDULES\\6.2 Schedules\\Doors\\SFS-COX-01-SH-AR91XX02-%s to %s CHANGE.xlsx'%(rev1,rev2))